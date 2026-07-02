# -*- coding: utf-8 -*-
"""EC매출 대시보드 — 매출현황/연간 결과를 웹에서 보기. 실행: streamlit run ec_dashboard.py"""
import io
import os
import glob
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="EC매출 대시보드", layout="wide", page_icon="📊")
st.markdown("""
<style>
.block-container {padding-top:1.6rem;}
div[data-testid="stMetric"]{background:#fff;border:1px solid #e6e6e6;border-radius:12px;
  padding:12px 16px;box-shadow:0 1px 3px rgba(0,0,0,.04);}
div[data-testid="stMetricLabel"]{font-size:.82rem;color:#666;}
div[data-testid="stMetricValue"]{font-size:1.4rem;}
h3{border-left:4px solid #2563eb;padding-left:10px;margin-top:1.4rem;}
h5{margin-bottom:.3rem;color:#374151;}
</style>
""", unsafe_allow_html=True)

_title_ph = st.empty()
_title_ph.title("📊 EC 매출 요약")


def _check_password():
    """비밀번호 게이트. secrets.toml의 password가 있으면 우선, 없으면 기본값."""
    if st.session_state.get("auth_ok"):
        return
    try:
        correct = st.secrets["password"]
    except Exception:
        correct = "ecearning"   # ← 기본 비밀번호 (원하는 값으로 바꿔 쓰기)
    pw = st.text_input("🔒 비밀번호를 입력하세요", type="password")
    if pw and pw == correct:
        st.session_state["auth_ok"] = True
        try:
            st.rerun()
        except AttributeError:
            st.experimental_rerun()
    if pw and pw != correct:
        st.error("비밀번호가 틀렸습니다.")
    st.stop()


_check_password()


def _find_latest():
    """스크립트 폴더 / data 폴더에서 최신 EC매출 파일 자동 탐색."""
    try:
        base = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base = os.getcwd()
    pats = []
    for d in [base, os.path.join(base, "data")]:
        pats += glob.glob(os.path.join(d, "EC매출*.xlsx"))
        pats += glob.glob(os.path.join(d, "EC매출*.xlsm"))
    pats = [p for p in pats if not os.path.basename(p).startswith("~")]
    return max(pats, key=os.path.getmtime) if pats else None


uploaded = st.file_uploader("EC매출 결과 파일 (.xlsx) — 안 올리면 폴더의 최신 파일 자동 사용", type=["xlsx", "xlsm"])
_auto = _find_latest()
if uploaded is not None:
    _data = uploaded.getvalue()
elif _auto:
    with open(_auto, "rb") as _f:
        _data = _f.read()
    st.caption(f"📂 자동 로드: {os.path.basename(_auto)}  (파일을 안 올리면 폴더의 최신 파일을 씁니다)")
else:
    st.info("EC매출 결과 파일을 올리거나, 이 스크립트 폴더(또는 data 폴더)에 EC매출*.xlsx 파일을 두세요. "
            "(엑셀에서 한 번 열어 계산된 파일 권장)")
    st.stop()


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def agg_raw(raw, img_map):
    """RAW(라인명 포함) → 병행/공식(pc), TOP10 브랜드, TOP30 상품(병행/공식). 필터 재집계용."""
    g = raw.groupby("분류")[["매출", "이익"]].sum()
    pc = {}
    for cls in ["병행", "공식"]:
        if cls in g.index:
            s = float(g.loc[cls, "매출"])
            p = float(g.loc[cls, "이익"])
            pc[cls] = {"매출": s, "이익": p, "수익율": p / s if s else 0.0}
        else:
            pc[cls] = {"매출": 0.0, "이익": 0.0, "수익율": 0.0}
    total_sales = float(raw["매출"].sum())
    top_brand = (raw.groupby("브랜드").agg(수량=("수량", "sum"), 매출=("매출", "sum"), 이익=("이익", "sum"))
                 .sort_values("매출", ascending=False).head(10))
    top_brand["객단가"] = (top_brand["매출"] / top_brand["수량"].replace(0, pd.NA)).fillna(0)
    top_brand["수익율"] = (top_brand["이익"] / top_brand["매출"]).fillna(0)
    top_brand["매출비중"] = (top_brand["매출"] / total_sales) if total_sales else 0.0

    def _build_top(rs):
        if not len(rs):
            return pd.DataFrame(columns=["매출", "이익", "수량", "브랜드", "카테고리", "이익율", "이미지"])
        tp = (rs.groupby("라인명")
              .agg(매출=("매출", "sum"), 이익=("이익", "sum"), 수량=("수량", "sum"),
                   브랜드=("브랜드", "first"), 카테고리=("카테고리", "first"))
              .sort_values("매출", ascending=False).head(30))
        tp["이익율"] = (tp["이익"] / tp["매출"]).fillna(0)
        _ms = rs.groupby(["라인명", "상품"])["매출"].sum().reset_index()
        _rep = _ms.loc[_ms.groupby("라인명")["매출"].idxmax()].set_index("라인명")["상품"]

        def _li(ln):
            u = img_map.get(str(ln).strip(), "")
            if u:
                return u
            return img_map.get(str(_rep.get(ln, "")).strip(), "")
        tp["이미지"] = [_li(ln) for ln in tp.index]
        return tp

    return pc, top_brand, _build_top(raw[raw["분류"] == "병행"]), _build_top(raw[raw["분류"] == "공식"])


@st.cache_data(show_spinner="파일 읽는 중...")
def load(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["매출현황"]
    month = int(_num(ws["B1"].value)) or 1
    prev = month - 1 if month > 1 else 12
    prev2 = month - 2 if month > 2 else month + 10
    v1 = _num(ws["V1"].value)   # 경과일
    v2 = _num(ws["V2"].value)   # 총일수
    day = int(v1) if v1 else 0
    is_last = (v2 > 0 and v1 >= v2)

    def box(r):
        tgt = ws.cell(r, 37).value   # AK 목표
        ach = ws.cell(r, 38).value   # AL 달성율
        return {"매출": _num(ws.cell(r, 34).value), "이익": _num(ws.cell(r, 35).value),
                "수익율": _num(ws.cell(r, 36).value),
                "목표": float(tgt) if isinstance(tgt, (int, float)) else None,
                "달성율": float(ach) if isinstance(ach, (int, float)) else None}
    cur_box, prev_box, prev2_box = box(8), box(16), box(24)

    cur_extra = {"목표": _num(ws.cell(8, 37).value), "달성율": _num(ws.cell(8, 38).value),
                 "예상": _num(ws.cell(8, 41).value), "예상신장율": _num(ws.cell(8, 42).value)}

    # 카테고리별 매출 (AG28:AK38 → 가방~용품+총합). AG=33
    cat_rows = []
    for r in range(30, 39):
        nm = ws.cell(r, 33).value
        if nm is None:
            continue
        cat_rows.append({"구분": nm, "매출": _num(ws.cell(r, 34).value),
                         "이익": _num(ws.cell(r, 35).value), "수익율": _num(ws.cell(r, 36).value),
                         "재고원가": _num(ws.cell(r, 37).value)})
    cat_sales = pd.DataFrame(cat_rows)

    # 쇼핑몰별 (전년 → 당월 → 전월 → 예상 → 대비)
    rows = []
    for r in range(6, 300):
        no = ws.cell(r, 2).value
        mall = ws.cell(r, 4).value
        if no is None and mall is None:
            break
        if mall is None:
            continue
        rows.append({
            "쇼핑몰": mall,
            "전년 수량": _num(ws.cell(r, 5).value), "전년 매출": _num(ws.cell(r, 6).value),
            "전년 이익": _num(ws.cell(r, 8).value), "전년 수익율": _num(ws.cell(r, 9).value),
            "당월 수량": _num(ws.cell(r, 10).value), "당월 매출": _num(ws.cell(r, 12).value),
            "목표": _num(ws.cell(r, 11).value), "당월 이익": _num(ws.cell(r, 14).value),
            "당월 수익율": _num(ws.cell(r, 15).value), "달성율": _num(ws.cell(r, 16).value),
            "전월 수량": _num(ws.cell(r, 17).value), "전월 매출": _num(ws.cell(r, 18).value),
            "전월 이익": _num(ws.cell(r, 20).value), "전월 수익율": _num(ws.cell(r, 21).value),
            "예상마감": _num(ws.cell(r, 22).value),
            "예상달성율": (_num(ws.cell(r, 22).value) / _num(ws.cell(r, 11).value))
                          if _num(ws.cell(r, 11).value) else None,
            "전년대비": _num(ws.cell(r, 29).value), "전월대비": _num(ws.cell(r, 31).value),
        })
    df = pd.DataFrame(rows)

    wy = wb["연간매출및 목표"]
    months, targets, actuals = [], [], []
    col = 8
    for m in range(1, 13):
        # 병행 합계행(6) + 공식 합계행(57)
        targets.append(_num(wy.cell(6, col).value) + _num(wy.cell(57, col).value))
        actuals.append(_num(wy.cell(6, col + 1).value) + _num(wy.cell(57, col + 1).value))
        months.append(f"{m}월")
        col += 3
    yearly = pd.DataFrame({"월": months, "목표": targets, "실제": actuals})

    # 카테고리분류: K(상품 식별자)→L(이미지URL), AB(모델명)→S(라인명)
    #  K:L은 S~AU(재고)와 행 정렬이 다른 독립 매핑이라 K로만 이미지 매칭.
    #  라인명(S)은 모델명(AB)에서 사이즈 표기 (40)(42) 뗀 것 → 사이즈 변형 묶음.
    img_map = {}
    model2line = {}
    try:
        wk = wb["카테고리분류"]
        for row in wk.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 28:
                continue
            k = row[10]; l = row[11]; s = row[18]; ab = row[27]
            if k and l and str(k).strip() and str(l).strip():
                img_map[str(k).strip()] = str(l).strip()
            if ab and s and str(ab).strip() and str(s).strip():
                model2line[str(ab).strip()] = str(s).strip()
    except Exception:
        pass

    # 병행/공식 + TOP (RAW: C=분류, F=브랜드, I=상품, J=수량, M=매출, V=이익, AA=카테고리)
    try:
        raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="RAW", header=1,
                            usecols=[2, 3, 5, 8, 9, 12, 21, 26])
        raw.columns = ["분류", "쇼핑몰", "브랜드", "상품", "수량", "매출", "이익", "카테고리"]
        for cc in ["수량", "매출", "이익"]:
            raw[cc] = pd.to_numeric(raw[cc], errors="coerce").fillna(0)
        # RAW 모델 → 라인명(S) 변환 (사이즈 변형 묶음)
        raw["라인명"] = raw["상품"].astype(str).str.strip().map(model2line)
        raw["라인명"] = raw["라인명"].fillna(raw["상품"].astype(str).str.strip())
        pc, top_brand, top_prod_bh, top_prod_gs = agg_raw(raw, img_map)
    except Exception:
        raw = pd.DataFrame(columns=["분류", "쇼핑몰", "브랜드", "상품", "수량", "매출", "이익", "카테고리", "라인명"])
        pc = {"병행": {"매출": 0.0, "이익": 0.0, "수익율": 0.0},
              "공식": {"매출": 0.0, "이익": 0.0, "수익율": 0.0}}
        top_brand = pd.DataFrame(columns=["수량", "매출", "이익", "객단가", "수익율", "매출비중"])
        top_prod_bh = pd.DataFrame(columns=["매출", "이익", "수량", "브랜드", "카테고리", "이익율"])
        top_prod_gs = pd.DataFrame(columns=["매출", "이익", "수량", "브랜드", "카테고리", "이익율"])

    # 병행/공식 당월 목표/달성율 (연간매출및 목표: 병행/공식 블록 합계행, 당월 목표매출 열)
    try:
        wy = wb["연간매출및 목표"]
        tcol = 8 + (month - 1) * 3   # 1월=8열, 이후 3칸 간격 → 당월 목표매출 열
        blk = {}
        for r in range(1, 160):
            v = wy.cell(r, 2).value
            if v == "병행":
                blk["병행"] = r
            elif v == "공식":
                blk["공식"] = r
        for label, sr in blk.items():
            for rr in range(sr, sr + 6):       # 블록 시작~ 헤더(3열에 '쇼') 찾아 다음 행=합계
                c3 = wy.cell(rr, 3).value
                if c3 and "쇼" in str(c3):
                    tgt = _num(wy.cell(rr + 1, tcol).value)
                    pc[label]["목표"] = tgt
                    pc[label]["달성율"] = (pc[label]["매출"] / tgt) if tgt else 0.0
                    break
    except Exception:
        pass
    for label in ["병행", "공식"]:
        pc[label].setdefault("목표", 0.0)
        pc[label].setdefault("달성율", 0.0)

    # 쇼핑몰 합계행 (매출현황 5행 = 엑셀 계산된 합계)
    def s5(cc):
        return _num(ws.cell(5, cc).value)
    total_row = {
        "쇼핑몰": "합계",
        "전년 수량": s5(5), "전년 매출": s5(6), "전년 이익": s5(8), "전년 수익율": s5(9),
        "당월 수량": s5(10), "당월 매출": s5(12), "목표": s5(11), "당월 이익": s5(14),
        "당월 수익율": s5(15), "달성율": s5(16),
        "전월 수량": s5(17), "전월 매출": s5(18), "전월 이익": s5(20), "전월 수익율": s5(21),
        "예상마감": s5(22), "예상달성율": (s5(22) / s5(11)) if s5(11) else None,
        "전년대비": s5(29), "전월대비": s5(31),
    }

    return (month, prev, prev2, day, is_last, cur_box, prev_box, prev2_box,
            cur_extra, cat_sales, df, yearly, total_row, pc, top_brand, top_prod_bh, top_prod_gs,
            raw, img_map)


try:
    (month, prev, prev2, day, is_last, cur_box, prev_box, prev2_box,
     cur_extra, cat_sales, df, yearly, total_row, pc, top_brand, top_prod_bh, top_prod_gs,
     raw, img_map) = load(_data)
except Exception as e:
    st.error(f"파일을 읽지 못했습니다 (매출현황/연간 시트 확인): {e}")
    st.stop()

_title_ph.title(f"📊 EC {month}월 매출 요약")

# ── 사이드바 필터 (쇼핑몰 / 카테고리) ──────────────────
_malls = sorted([m for m in raw["쇼핑몰"].dropna().astype(str).unique() if m.strip()])
_cats = sorted([c for c in raw["카테고리"].dropna().astype(str).unique() if c.strip()])
with st.sidebar:
    st.header("🔎 필터")
    st.caption("TOP 브랜드 · TOP 상품 섹션에 적용")
    sel_mall = st.multiselect("쇼핑몰", _malls)
    sel_cat = st.multiselect("카테고리", _cats)
raw_f = raw
if sel_mall:
    raw_f = raw_f[raw_f["쇼핑몰"].astype(str).isin(sel_mall)]
if sel_cat:
    raw_f = raw_f[raw_f["카테고리"].astype(str).isin(sel_cat)]
if sel_mall or sel_cat:
    _, top_brand, top_prod_bh, top_prod_gs = agg_raw(raw_f, img_map)
    _bits = ([f"쇼핑몰 {len(sel_mall)}개"] if sel_mall else []) + ([f"카테고리 {len(sel_cat)}개"] if sel_cat else [])
    st.info("🔎 필터: " + " / ".join(_bits) + "  — TOP 브랜드·TOP 상품에만 반영")

if cur_box["매출"] == 0 and df["당월 매출"].sum() == 0:
    st.warning("숫자가 전부 0이에요. EC매출 결과는 **엑셀에서 한 번 열어 계산된 뒤 저장**해야 값이 채워집니다.")

cur_label = f"{month}월" if is_last else f"{month}월(~{month}/{day})"

# ── 종합 ───────────────────────────────────────────────
title = f"{month}월 매출" if is_last else f"~{month}/{day} 까지 매출"
st.subheader(f"📊 {title}")
c = st.columns(4 if is_last else 5)
c[0].metric("매출", f"{cur_box['매출']/1e8:,.2f}억")
c[1].metric("이익", f"{cur_box['이익']/1e8:,.2f}억", f"수익율 {cur_box['수익율']:.1%}")
c[2].metric("목표", f"{cur_extra['목표']/1e8:,.2f}억")
c[3].metric("달성율", f"{cur_extra['달성율']:.1%}")
if not is_last:
    _fc_ach = (cur_extra['예상'] / cur_extra['목표']) if cur_extra['목표'] else 0
    c[4].metric("예상마감", f"{cur_extra['예상']/1e8:,.2f}억", f"달성율 {_fc_ach:.1%}")

# ── 월별(작게) + 카테고리별 나란히 ─────────────────────
c_left, c_right = st.columns([1, 1.35])
with c_left:
    st.markdown("##### 📅 월별 매출")
    mdf = pd.DataFrame([
        {"월": f"{prev2}월", "목표": prev2_box["목표"], "매출": prev2_box["매출"],
         "이익": prev2_box["이익"], "수익율": prev2_box["수익율"], "달성율": prev2_box["달성율"]},
        {"월": f"{prev}월", "목표": prev_box["목표"], "매출": prev_box["매출"],
         "이익": prev_box["이익"], "수익율": prev_box["수익율"], "달성율": prev_box["달성율"]},
        {"월": cur_label, "목표": cur_box["목표"], "매출": cur_box["매출"],
         "이익": cur_box["이익"], "수익율": cur_box["수익율"], "달성율": cur_box["달성율"]},
    ])
    st.dataframe(
        mdf.style.format({"매출": "{:,.0f}", "이익": "{:,.0f}", "수익율": "{:.1%}",
                          "목표": "{:,.0f}", "달성율": "{:.1%}"}, na_rep="-"),
        hide_index=True, use_container_width=True,
    )
    st.markdown("##### 🔀 병행 / 공식 (당월)")
    pcdf = pd.DataFrame([
        {"구분": "병행", "목표": pc["병행"]["목표"], "매출": pc["병행"]["매출"],
         "이익": pc["병행"]["이익"], "수익율": pc["병행"]["수익율"], "달성율": pc["병행"]["달성율"]},
        {"구분": "공식", "목표": pc["공식"]["목표"], "매출": pc["공식"]["매출"],
         "이익": pc["공식"]["이익"], "수익율": pc["공식"]["수익율"], "달성율": pc["공식"]["달성율"]},
    ])
    st.dataframe(
        pcdf.style.format({"목표": "{:,.0f}", "매출": "{:,.0f}", "이익": "{:,.0f}",
                           "수익율": "{:.1%}", "달성율": "{:.1%}"}, na_rep="-"),
        hide_index=True, use_container_width=True,
    )
with c_right:
    st.markdown("##### 🗂 카테고리별 매출")
    st.dataframe(
        cat_sales.style.format({
            "매출": "{:,.0f}", "이익": "{:,.0f}", "수익율": "{:.1%}", "재고원가": "{:,.0f}",
        }),
        hide_index=True, use_container_width=True, height=350,
    )

# ── 목표 (연간 월별 + 누적 달성율) ────────────────────
st.subheader("📈 목표")
year_t = float(yearly["목표"].sum())            # 1년 전체 목표
cum_a = float(yearly["실제"][:month].sum())     # 누적 실제
ach = cum_a / year_t if year_t else 0.0
mc = st.columns(3)
mc[0].metric("연간 목표", f"{year_t/1e8:,.1f}억")
mc[1].metric("누적 실제", f"{cum_a/1e8:,.1f}억")
mc[2].metric("목표달성율", f"{ach:.1%}")
fig = go.Figure()
fig.add_bar(x=yearly["월"], y=yearly["목표"], name="목표", marker_color="#cbd5e1")
fig.add_bar(x=yearly["월"], y=yearly["실제"], name="실제", marker_color="#2563eb")
fig.update_layout(barmode="group", height=360, margin=dict(t=20, b=20, l=20, r=20),
                  legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
                  yaxis=dict(tickformat=","))
st.plotly_chart(fig, use_container_width=True)

# ── 쇼핑몰별 표 (원본 색 + 순서) ───────────────────────
st.subheader(f"🏬 쇼핑몰별 ({cur_label})")
_ca, _cb = st.columns(2)
hide_zero = _ca.checkbox("매출 0인 쇼핑몰 숨기기", value=True)
show_detail = _cb.checkbox("이익/수량 보기 (전년/당월/전월)", value=False)
view = df[df["당월 매출"] > 0] if hide_zero else df
view = view.sort_values("당월 매출", ascending=False).reset_index(drop=True)
# 맨 위에 합계 행
view = pd.concat([pd.DataFrame([total_row]), view], ignore_index=True)
st.caption(f"{len(view)-1}개 쇼핑몰 · 당월 매출 큰 순  (⬛합계 🟦당월 🟨전월 ⬜전년)")

LY = ["전년 수량", "전년 매출", "전년 이익", "전년 수익율"]
CUR = ["당월 수량", "목표", "당월 매출", "당월 이익", "당월 수익율", "달성율"]
PV = ["전월 수량", "전월 매출", "전월 이익", "전월 수익율"]
money = ["전년 매출", "전년 이익", "당월 매출", "목표", "당월 이익", "전월 매출", "전월 이익", "예상마감"]
qty = ["전년 수량", "당월 수량", "전월 수량"]
pct = ["전년 수익율", "당월 수익율", "전월 수익율", "달성율", "예상달성율"]
gpct = ["전년대비", "전월대비"]

# 표 컬럼 표시 순서 (CUR 순서 반영: 목표 → 당월 매출) + 이익·수량 기본 숨김
_drop = [] if show_detail else ["전년 수량", "전년 이익", "당월 수량", "당월 이익", "전월 수량", "전월 이익"]
_order = ["쇼핑몰"] + LY + CUR + PV + ([] if is_last else ["예상마감", "예상달성율"]) + gpct
view = view[[c for c in _order if c in view.columns and c not in _drop]]

def bg(data):
    s = pd.DataFrame("", index=data.index, columns=data.columns)
    for cc in CUR:
        if cc in s.columns:
            s[cc] = "background-color:#DCE6F7"
    for cc in PV:
        if cc in s.columns:
            s[cc] = "background-color:#FCF4DC"
    for cc in LY:
        if cc in s.columns:
            s[cc] = "background-color:#F4F4F4"
    if len(data) > 0:                       # 합계 행(0) 강조
        s.iloc[0] = ["background-color:#D1D5DB; font-weight:700"] * len(data.columns)
    return s

def neg_red(v):
    try:
        return "color:#c00; font-weight:600" if float(v) < 0 else ""
    except (TypeError, ValueError):
        return ""

fmt = {**{cc: "{:,.0f}" for cc in money + qty},
       **{cc: "{:.1%}" for cc in pct},
       **{cc: "{:+.1%}" for cc in gpct}}
sty = view.style.apply(bg, axis=None).format(fmt)
sty = sty.set_properties(**{"text-align": "center"})
sty = sty.set_table_styles([
    {"selector": "", "props": [("border-collapse", "separate"), ("border-spacing", "0")]},
    {"selector": "th", "props": [("text-align", "center"), ("background-color", "#eef2f7"),
                                 ("padding", "5px 8px"), ("white-space", "nowrap"),
                                 ("position", "sticky"), ("top", "33px"), ("z-index", "3")]},
    {"selector": "td", "props": [("text-align", "center"), ("padding", "5px 8px"),
                                 ("white-space", "nowrap")]},
    {"selector": "tbody tr:first-child td",
     "props": [("position", "sticky"), ("top", "66px"), ("z-index", "2")]},
])
try:
    sty = sty.map(neg_red, subset=gpct)
except AttributeError:
    sty = sty.applymap(neg_red, subset=gpct)
try:
    html = sty.hide(axis="index").to_html()
except (AttributeError, TypeError):
    html = sty.hide_index().to_html()
# 그룹 헤더 행 (전년/당월/전월/마감) — 2단 헤더
_GROUP = {
    "전년 수량": "전년", "전년 매출": "전년", "전년 이익": "전년", "전년 수익율": "전년",
    "당월 수량": "당월", "목표": "당월", "당월 매출": "당월", "당월 이익": "당월",
    "당월 수익율": "당월", "달성율": "당월",
    "전월 수량": "전월", "전월 매출": "전월", "전월 이익": "전월", "전월 수익율": "전월",
    "예상마감": "마감", "예상달성율": "마감", "전년대비": "마감", "전월대비": "마감",
}
_gcolors = {"전년": "#E5E7EB", "당월": "#C7D7F0", "전월": "#F5E9C8", "마감": "#E8EDF3"}
_gc = {}
for _c in view.columns:
    _g = _GROUP.get(_c, "")
    _gc[_g] = _gc.get(_g, 0) + 1
_gcells = ""
for _g, _n in _gc.items():
    if _g == "":
        _gcells += ('<th style="background-color:#eef2f7;position:sticky;top:0;'
                    'z-index:5;border:1px solid #e5e7eb;"></th>')
    else:
        _gcells += (f'<th colspan="{_n}" style="text-align:center;font-weight:700;'
                    f'background-color:{_gcolors.get(_g, "#eef2f7")};padding:6px 8px;'
                    f'position:sticky;top:0;z-index:5;border:1px solid #e5e7eb;">{_g}</th>')
html = html.replace("<thead>", f'<thead><tr>{_gcells}</tr>', 1)
st.markdown(
    f'<div style="overflow:auto; max-height:600px; font-size:13px; '
    f'border:1px solid #e5e7eb; border-radius:8px">{html}</div>',
    unsafe_allow_html=True,
)

# ── TOP10 브랜드 (도넛 + 표) ───────────────────────────
st.subheader("🏷 TOP 10 브랜드")
if len(top_brand):
    bc1, bc2 = st.columns([1, 1.35])
    with bc1:
        fig = go.Figure(go.Pie(
            labels=list(top_brand.index), values=list(top_brand["매출"]), hole=0.55,
            textinfo="label+percent", textposition="inside", sort=False,
            insidetextorientation="horizontal",
        ))
        fig.update_layout(height=400, margin=dict(t=10, b=10, l=10, r=10), showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
    with bc2:
        bt = top_brand.reset_index()[["브랜드", "수량", "매출", "객단가", "수익율", "매출비중"]]
        bt.insert(0, "#", range(1, len(bt) + 1))

        def neg_red_b(v):
            try:
                return "color:#c00; font-weight:600" if float(v) < 0 else ""
            except (TypeError, ValueError):
                return ""
        bsty = bt.style.format({
            "수량": "{:,.0f}", "매출": "{:,.0f}", "객단가": "{:,.0f}",
            "수익율": "{:.1%}", "매출비중": "{:.1%}",
        }).set_properties(**{"text-align": "center"})
        bsty = bsty.set_table_styles([
            {"selector": "th", "props": [("text-align", "center"), ("background-color", "#eef2f7"),
                                         ("padding", "6px 10px"), ("white-space", "nowrap")]},
            {"selector": "td", "props": [("text-align", "center"), ("padding", "6px 10px"),
                                         ("white-space", "nowrap")]},
        ])
        try:
            bsty = bsty.map(neg_red_b, subset=["수익율"])
        except AttributeError:
            bsty = bsty.applymap(neg_red_b, subset=["수익율"])
        try:
            bhtml = bsty.hide(axis="index").to_html()
        except (AttributeError, TypeError):
            bhtml = bsty.hide_index().to_html()
        st.markdown(
            '<div style="overflow:auto; font-size:13px; border:1px solid #e5e7eb; '
            'border-radius:8px">' + bhtml + '</div>',
            unsafe_allow_html=True,
        )
else:
    st.caption("데이터 없음")

# ── TOP30 상품 (병행 / 공식 탭) ────────────────────────
st.subheader("📦 TOP 30 상품")


def _render_top30(tp):
    if not len(tp):
        st.caption("데이터 없음")
        return
    pcards = ""
    for i, (model, r) in enumerate(tp.iterrows(), 1):
        url = r["이미지"] if "이미지" in tp.columns else ""
        if url:
            img_html = (
                f'<img src="{url}" style="width:64px;height:64px;object-fit:cover;'
                'border-radius:8px;border:1px solid #eef2f7;background:#f1f5f9;flex:0 0 auto;">'
            )
        else:
            img_html = (
                '<div style="width:64px;height:64px;background:#f1f5f9;border:1px solid #eef2f7;'
                'border-radius:8px;display:flex;align-items:center;justify-content:center;'
                'color:#cbd5e1;font-size:22px;flex:0 0 auto;">📦</div>'
            )
        pcards += (
            '<div style="display:flex;gap:10px;align-items:center;padding:10px;'
            'border:1px solid #eef2f7;border-radius:10px;background:#fff;">'
            + img_html +
            '<div style="min-width:0;flex:1;">'
            f'<div style="font-size:10px;color:#94a3b8;white-space:nowrap;overflow:hidden;'
            f'text-overflow:ellipsis;">#{i} · {r.브랜드} · {r.카테고리}</div>'
            f'<div style="font-size:13px;font-weight:600;color:#0f172a;white-space:nowrap;'
            f'overflow:hidden;text-overflow:ellipsis;">{model}</div>'
            f'<div style="font-size:12px;color:#0f172a;">{r.매출:,.0f}'
            f'<span style="color:#64748b;"> · {r.이익율:.1%} · {r.수량:,.0f}개</span></div>'
            '</div></div>'
        )
    st.markdown(
        '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;">'
        + pcards + '</div>',
        unsafe_allow_html=True,
    )
    if "이미지" not in tp.columns or not any(tp["이미지"]):
        st.caption("📦 카테고리분류 L(이미지URL) 채우면 자동으로 떠요 (모델명 또는 라인명 기준)")


_tab_bh, _tab_gs = st.tabs(["🟦 병행 TOP 30", "🟩 공식 TOP 30"])
with _tab_bh:
    _render_top30(top_prod_bh)
with _tab_gs:
    _render_top30(top_prod_gs)

csv = view.to_csv(index=False).encode("utf-8-sig")
st.download_button("⬇ 쇼핑몰별 표 CSV로 받기", csv,
                   file_name=f"EC매출_{month}월_쇼핑몰별.csv", mime="text/csv")
